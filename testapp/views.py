from django.shortcuts import render
import openpyxl
import os
import json
from django.http import JsonResponse

def index(request):
    return render(request, 'index.html')

def get_questions(request):
    questions = load_questions()
    return JsonResponse(questions, safe=False)

def load_questions():
    questions = []
    # Make path relative to project directory
    import os
    from django.conf import settings
    excel_path = os.path.join(settings.BASE_DIR, "Pedagogika Psixologiya yakuniy savollar bazasi.xlsx")
    
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        
        # Start from the second row (index 1)
        for row in range(2, sheet.max_row + 1):
            question_num = sheet.cell(row=row, column=1).value
            question_text = sheet.cell(row=row, column=2).value
            
            # Skip empty rows
            if not question_text:
                continue
            
            # Get answer options
            options = {}
            options["A"] = sheet.cell(row=row, column=3).value
            options["B"] = sheet.cell(row=row, column=4).value
            options["C"] = sheet.cell(row=row, column=5).value
            options["D"] = sheet.cell(row=row, column=6).value
            
            # Get correct answer text from column G
            correct_answer_text = sheet.cell(row=row, column=7).value
            
            # Find which option matches the correct answer text
            correct_option = None
            for option_key, option_text in options.items():
                if option_text and correct_answer_text and option_text.strip() == correct_answer_text.strip():
                    correct_option = option_key
                    break
                    
            # If no match found, use the first option as fallback (shouldn't happen with proper data)
            if not correct_option and options["A"]:
                correct_option = "A"
                print(f"Warning: No matching option found for answer '{correct_answer_text}' in question {question_num}")
            
            # Add question to list
            questions.append({
                "number": question_num,
                "text": question_text,
                "options": options,
                "correct_answer": correct_option,
                "correct_text": correct_answer_text
            })
        
        return questions
    except Exception as e:
        print(f"Error loading questions: {e}")
        return []
